import shutil
import pandas as pd
import logging
import os
from openpyxl import load_workbook
import glob

# Configuration for logging

logging.basicConfig(
    filename="debug.log",
    level=logging.DEBUG,
    format="[%(asctime)s] %(lineno)d [%(levelname)s   :] - %(message)s",
    filemode="w",
    datefmt="%H:%M:%S",
)


def checkPromoter(promoters):
    try:
        logging.info("Checking Promoter Value and determing Quality")
        if promoters > 200:
            return "Good"
        elif promoters == 200:
            return "Average"
        else:
            return "Bad"
    except TypeError as err:
        logging.debug("Promoter Error: {}".format(err))
        return


def checkPassive(passive):
    try:
        logging.info("Checking Passive Value and determing Quality")
        if passive > 100:
            return "Good"
        elif passive == 100:
            return "Average"
        else:
            return "Bad"
    except TypeError as err:
        logging.debug("Passive Error: {}".format(err))
        return


def checkDecratctor(decarator):
    try:
        logging.info("Checking Decarator Value and determing Quality")
        if decarator > 100:
            return "Good"
        elif decarator == 100:
            return "Average"
        else:
            return "Bad"
    except TypeError as err:
        logging.debug("Decrators Error: {}".format(err))
        return


def getMonthAndYear():
    try:
        for i in range(len(df.columns[1:])):
            if "-" not in str(df.columns[i + 1]):
                columnName.append("{}-{}".format(str(df.columns[i + 1]).lower(), year))
            else:
                Year = str(df.columns[i + 1]).split("-")[0]
                Month = month_labels[int(str(df.columns[i + 1]).split("-")[1])]
                columnName.append("{}-{}".format(Month, Year))
        df.columns = columnName
        return
    except TypeError as err:
        logging.debug("Getting Month and Year Error: {}".format(err))
        return


def formatTable(df):
    try:
        logging.info("Formating Table...")
        df = df[df.columns[0 : numberOfMonthInYear + 1]]
        df = df[2:8:2]
        df.index = ["Promoter", "Passive", "Decratctor"]
        return df
    except BaseException as err:
        logging.debug("Error Formating Table \n {}".format(err))
        return df


def getUserData():
    try:
        Date = "{}-{}".format(month, year)
        promoter = checkPromoter(df.loc["Promoter", Date])
        passive = checkPassive(df.loc["Passive", Date])
        decratctor = checkDecratctor(df.loc["Decratctor", Date])

        logging.info("\n\n\n{}".format(fileName))
        logging.info("Promoter = {}".format(promoter))
        logging.info("Passive = {}".format(passive))
        logging.info("Decratctor = {}\n\n\n".format(decratctor))
    except TypeError as err:
        logging.debug("Error Getting User Data: \n {}".format(err))
        return


def moveToArchive(fileName):
    try:
        logging.info("Moving File {} to Archive Folder".format(fileName))
        f = fileName.split("\\")[-1]
        shutil.move(path + "\\" + f, path + "\\Archive\\" + f)
    except BaseException as err:
        logging.debug("Error Moving File into Error Folder. {}".format(err))


def moveToError(fileName):
    try:
        logging.info("Moving File {} to Error Folder".format(fileName))
        f = fileName.split("\\")[-1]
        shutil.move(path + "\\" + f, path + "\\Error\\" + f)
    except BaseException as err:
        logging.debug("Error Moving File into Error Folder. {}".format(err))


def checkSheetName(fileName):
    try:
        logging.info("Checking if SheetName Exist")
        result = True
        wb = load_workbook(fileName, read_only=True)
        for i in range(len(sheetNameList)):
            if sheetNameList[i] not in wb.sheetnames:
                result = False
        wb.close()
        return result
    except BaseException as err:
        logging.debug("Error with checking sheet name: {}".format(err))


def checkFileName(fileName):
    logging.info("Checking if File Name Meet Criteria")
    monthExist = False
    fileNameVerification = fileName.split("_")
    try:
        if len(fileNameVerification) != 5:
            return False
        if fileNameVerification[0] != "expedia":
            return False
        if fileNameVerification[1] != "report":
            return False
        if fileNameVerification[2] != "monthly":
            return False
        for i in range(numberOfMonthInYear):
            if fileNameVerification[3] == month_labels[i + 1]:
                monthExist = True
        if monthExist is False:
            return False
        return True
    except BaseException as err:
        logging.DEBUG("Error Checking File Name is valid: {}".format(err))
        return False


def getFileList():
    result = []
    ErrorPath = path + "\\" + "Error"
    ArchivePath = path + "\\" + "Archive"

    try:
        logging.info("Attempting to get file list ...")
        errfiles = glob.glob(os.path.join(ErrorPath, "*.xlsx"))
        for i in range(len(errfiles)):
            result.append((errfiles[i].split("\\")[-1:])[0])
        archivefiles = glob.glob(os.path.join(ArchivePath, "*.xlsx"))
        for i in range(len(archivefiles)):
            result.append((archivefiles[i].split("\\")[-1:])[0])
        return result
    except BaseException as err:
        logging.debug("Error getting file list: {}".format(err))


# Variables with values that will not be reassigned
logging.info("Initializing Global Variables ...")


# Variable used to convert datetime month to words
month_labels = {
    1: "january",
    2: "feburary",
    3: "march",
    4: "april",
    5: "may",
    6: "june",
    7: "july",
    8: "august",
    9: "september",
    10: "october",
    11: "november",
    12: "december",
}
numberOfMonthInYear = 12
path = os.getcwd()
fileNames = glob.glob(os.path.join(path, "*.xlsx"))
sheetName = "VOC Rolling MoM"
sheetNameList = [
    "Summary Rolling MoM",
    "VOC Rolling MoM",
    "Monthly Verbatim Statements",
]
decoder = "openpyxl"
logging.info("Global Variables Values Have Been Obtained")

# Variable with values that will be reassign throuhout the code
columnName = ["Date"]
fileList = []

fileList = getFileList()
try:
    for fileName in fileNames:
        logging.info("Checking if File {} exist within the file list".format(fileName))
        if (fileName.split("\\")[-1:])[0] in fileList:
            logging.info(
                "The following file cannot be processed, it exist within either the archive or error file \n {}".format(
                    fileName
                )
            )
            continue
        logging.info("File Checking Complete!")

        # Check if format of name is correct
        if checkFileName(fileName.split("\\")[-1:][0]) is False:
            moveToError(fileName)
            continue
        # Check if a certain Sheet Name exist
        if checkSheetName(fileName) is False:
            moveToError(fileName)
            continue
        month = fileName.split("_")[3]
        year = int((fileName.split("_")[4])[0:4])
        df = pd.read_excel(fileName, sheet_name=sheetName, engine=decoder)
        df = formatTable(df)
        getMonthAndYear()
        getUserData()
        moveToArchive(fileName)
        fileList.append(fileName)
        columnName = ["Date"]


except BaseException as err:
    logging.debug("Error: The following file {} does not exist".format(err))
    print("The program will now exit.")
    quit()
